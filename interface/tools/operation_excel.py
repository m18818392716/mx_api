# -*- coding:utf-8 -*-
# @Time    : 2019/3/27 下午3:51
# @Author  : Susanna Chen
# @Site    : 
# @File    : operation_excel.py
# @Software: PyCharm

import xlrd
# import xlwt
# from xlutils.copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, colors


class OperationExcel:
    def __init__(self, file_name=None, sheet_id=None):
        if file_name:
            self.file_name = file_name
            self.sheet_id = sheet_id
        else:
            self.file_name = '../dataconfig/case1.xlsx'
            self.sheet_id = 0
        self.data = self.get_data()

    # 获取sheets的内容
    def get_data(self):
        data = xlrd.open_workbook(self.file_name)
        tables = data.sheets()[self.sheet_id]
        return tables

    # 获取单元格的行数
    def get_lines(self):
        tables = self.data
        return tables.nrows

    # 获取某一个单元格的内容
    def get_cell_value(self, row, col):
        return self.data.cell_value(row, col)

    # 写入数据
    def write_value(self, row, col, value):
        '''
        写入excel数据
        row,col,value
        '''
        # style = "font: color-index red,bold on"
        # red_style = xlwt.easyxf(style)
        #
        # read_data = xlrd.open_workbook(self.file_name)
        # write_data = copy(read_data)
        # sheet_data = write_data.get_sheet(0)
        # if value == 'pass':
        #     sheet_data.write(row,col,value)
        # else:
        #     sheet_data.write(row, col, value,red_style)
        # write_data.save(self.file_name)

        # 使用openpyxl进行修改xlsx文件
        # 生成一个已存在的wookbook对象
        wb = load_workbook(self.file_name)
        sheet_data = wb.worksheets[0]
        # openpyxl中行与列的索引是从1开始！
        result_cell = sheet_data.cell(row + 1, col + 1, value)
        if value == 'failed':
            result_cell.font = Font('Arial', color=colors.RED, bold=True)
        wb.save(self.file_name)  # 保存替换

    # 根据对应的caseid 找到对应行的内容
    def get_rows_data(self, case_id):
        row_num = self.get_row_num(case_id)
        rows_data = self.get_row_values(row_num)
        return rows_data

    # 根据对应的caseid找到对应的行号
    def get_row_num(self, case_id):
        num = 0
        clols_data = self.get_cols_data()
        for col_data in clols_data:
            if str(case_id) in col_data:
                return num
            num = num + 1

    # 根据行号，找到该行的内容
    def get_row_values(self, row):
        tables = self.data
        row_data = tables.row_values(row)
        return row_data

    # 获取某一列的内容
    def get_cols_data(self, col_id=None):
        if col_id != None:
            cols = self.data.col_values(col_id)
        else:
            cols = self.data.col_values(0)
        return cols


if __name__ == '__main__':
    opers = OperationExcel()
    print(opers.get_cell_value(1, 2))
