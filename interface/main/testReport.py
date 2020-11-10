# coding:utf-8
import xlsxwriter, openpyxl, os, time
from tools.configData import configKV
from openpyxl import styles
# from report.collectReportData import reportData

class reportExcel(object):
    # lastrow=0
    def get_format(self, wd, option={}):
        return wd.add_format(option)

    # 设置居中样式
    def get_format_center(self, wd, num=1):
        return wd.add_format({'align': 'left', 'valign': 'vcenter', 'border': num})

    def set_border_(self, wd, num=1):
        return wd.add_format({}).set_border(num)

    # 写数据
    def _write_center(self, worksheet, cl, data, wd):
        return worksheet.write(cl, data, self.get_format_center(wd))

    # 总览表，饼图表示
    def overview_init(self, worksheet, workbook,
                      datapie):  # data = {"test_name": "MMS","test_sum": 110,"test_success": 80,"test_failed": 20,"test_not":10,'defect_rate':18,'total_serious':6,'total_medium':5,'total_commonly':8}
        # 设置列行的宽高
        worksheet.set_column("A:A", 15)  # 设置一列或者多列单元属性
        worksheet.set_column("B:B", 15)
        worksheet.set_column("C:C", 15)
        worksheet.set_column("D:D", 15)
        worksheet.set_column("E:E", 15)
        worksheet.set_column("F:F", 15)
        worksheet.set_column("G:G", 15)
        worksheet.set_column("H:H", 15)
        worksheet.set_column("I:I", 15)
        # worksheet.set_row(1, 30)
        # worksheet.set_row(2, 30)
        # worksheet.set_row(3, 30)
        define_format_H1 = self.get_format(workbook,
                                           {'bold': True, 'font_size': 18})  # 在工作表中创建一个新的格式对象来格式化单元格，实现加粗，字体18号大小
        define_format_H2 = self.get_format(workbook,
                                           {'bold': True, 'font_size': 14})  # 在工作表中创建一个新的格式对象来格式化单元格，实现加粗，字体14号大小
        define_format_H1.set_border(1)  # 边框加粗1
        define_format_H2.set_border(1)
        define_format_H1.set_align("center")
        define_format_H2.set_align("center")
        define_format_H2.set_bg_color("blue")  # 填充背景颜色蓝色
        define_format_H2.set_color("#ffffff")  # 填充单元格内容的颜色为白色
        worksheet.merge_range('A1:I1', '测试结果总概况', define_format_H1)
        worksheet.merge_range('A2:I2', '测试概括', define_format_H2)
        # worksheet.merge_range('A3:A6', '这里放图片', self.get_format_center(workbook))
        self._write_center(worksheet, "A3", '项目名称', workbook)
        self._write_center(worksheet, "B3", "总用例数", workbook)
        self._write_center(worksheet, "C3", "通过总数", workbook)
        self._write_center(worksheet, "D3", "失败总数", workbook)
        self._write_center(worksheet, "E3", "未测试总数", workbook)
        self._write_center(worksheet, "F3", '缺陷率(%)', workbook)
        self._write_center(worksheet, "G3", "严重缺陷", workbook)
        self._write_center(worksheet, "H3", "中等缺陷", workbook)
        self._write_center(worksheet, "I3", "一般缺陷", workbook)

        self._write_center(worksheet, "A4", datapie['test_name'], workbook)
        self._write_center(worksheet, "B4", datapie['test_sum'], workbook)
        self._write_center(worksheet, "C4", datapie['test_success'], workbook)
        self._write_center(worksheet, "D4", datapie['test_failed'], workbook)
        self._write_center(worksheet, "E4", datapie['test_not'], workbook)
        self._write_center(worksheet, "F4", datapie['defect_rate'], workbook)
        self._write_center(worksheet, "G4", datapie['total_serious'], workbook)
        self._write_center(worksheet, "H4", datapie['total_medium'], workbook)
        self._write_center(worksheet, "I4", datapie['total_commonly'], workbook)
        self.overall_pie(workbook, worksheet)
        self.bug_grade_pie(workbook, worksheet)

    # bug分布表，柱状图表示
    def distribution_init(self, worksheet, workbook,
                          datacolumn):  # {'load-exchange-rate':{'serious':3,'medium':5,'commonly':2,'percent':30},'classify':{'serious':3,'medium':5,'commonly':2,'percent':30}.....}
        define_format_H1 = self.get_format(workbook,
                                           {'bold': True, 'font_size': 18})  # 在工作表中创建一个新的格式对象来格式化单元格，实现加粗，字体18号大小
        define_format_H2 = self.get_format(workbook,
                                           {'bold': True, 'font_size': 14})  # 在工作表中创建一个新的格式对象来格式化单元格，实现加粗，字体14号大小
        define_format_H1.set_border(1)  # 边框加粗1
        define_format_H2.set_border(1)
        define_format_H1.set_align("center")
        define_format_H2.set_align("center")
        define_format_H2.set_bg_color("blue")  # 填充背景颜色蓝色
        define_format_H2.set_color("#ffffff")  # 填充单元格内容的颜色为白色
        worksheet.merge_range('A20:F20', '各个接口缺陷分布情况', define_format_H1)
        worksheet.merge_range('A21:F21', '缺陷分布', define_format_H2)
        self._write_center(worksheet, "A22", '接口名称', workbook)
        self._write_center(worksheet, "B22", '严重缺陷', workbook)
        self._write_center(worksheet, "C22", "中等缺陷", workbook)
        self._write_center(worksheet, "D22", "一般缺陷", workbook)
        self._write_center(worksheet, "E22", "缺陷总数", workbook)
        self._write_center(worksheet, "F22", "百分比(%)", workbook)
        keys_interface = datacolumn.keys()
        col = 22
        for interfaceName in keys_interface:
            col = col + 1
            self._write_center(worksheet, 'A' + str(col), interfaceName, workbook)
            self._write_center(worksheet, 'B' + str(col), datacolumn[interfaceName]['serious'], workbook)
            self._write_center(worksheet, 'C' + str(col), datacolumn[interfaceName]['medium'], workbook)
            self._write_center(worksheet, 'D' + str(col), datacolumn[interfaceName]['commonly'], workbook)
            self._write_center(worksheet, 'E' + str(col), datacolumn[interfaceName]['total'], workbook)
            self._write_center(worksheet, 'F' + str(col), datacolumn[interfaceName]['percent'], workbook)
        self.bug_grade_column(workbook, worksheet, col)
        # self.lastrow=col

    # 生成总体情况的饼形图
    def overall_pie(self, workbook, worksheet):
        chart1 = workbook.add_chart({'type': 'pie'})
        chart1.add_series({
            'name': '总体情况图',
            'categories': '=测试总况!$C$3:$E$3',
            'values': '=测试总况!$C$4:$E$4',
        })
        chart1.set_title({'name': '总体情况图'})
        chart1.set_style(10)
        chart1.set_size({'width': 400, 'height': 250})
        worksheet.insert_chart('A6', chart1)

        # 生成缺陷情况的饼形图

    def bug_grade_pie(self, workbook, worksheet):
        chart1 = workbook.add_chart({'type': 'pie'})
        chart1.add_series({
            'name': '各严重程度缺陷所占比例',
            'categories': '=测试总况!$G$3:$I$3',
            'values': '=测试总况!$G$4:$I$4',
        })
        chart1.set_title({'name': '各严重程度缺陷所占比例'})
        chart1.set_style(10)
        chart1.set_size({'width': 400, 'height': 250})
        worksheet.insert_chart('E6', chart1)

    # 柱状图统计各个接口各种严重程度的缺陷数
    def bug_grade_column(self, workbook, worksheet, lastrow):
        chart1 = workbook.add_chart({"type": "column"})
        chart1.add_series({
            "name": "=测试总况!$B$22",  # 图例项
            "categories": "=测试总况!$A$23:$A$" + str(lastrow),  # X轴 Item名称
            "values": "=测试总况!$B$23:$B$" + str(lastrow)  # X轴Item值
        })
        chart1.add_series({
            "name": "=测试总况!$C$22",
            "categories": "=测试总况!$A$23:$A$" + str(lastrow),
            "values": "=测试总况!$C$23:$C$" + str(lastrow)
        })
        chart1.add_series({
            "name": "=测试总况!$D$22",
            "categories": "=测试总况!$A$23:$A$" + str(lastrow),
            "values": "=测试总况!$D$23:$D$" + str(lastrow)
        })
        # 添加柱状图标题
        chart1.set_title({"name": "各接口缺陷数分布图"})
        # Y轴名称
        chart1.set_y_axis({"name": "缺陷数"})
        # X轴名称
        chart1.set_x_axis({"name": "缺陷严重等级"})
        # 图表样式
        chart1.set_style(11)
        if (lastrow < 29):
            lastrow = 28
        chart1.set_size({'width': 80 * (lastrow - 23), 'height': 388})
        worksheet.insert_chart('A' + str(lastrow), chart1, {'x_offset': (lastrow - 22) * 5, 'y_offset': 25})

    # 将filename1的sheetname1的内容复制到filename2的sheetname2中
    def replace_xls(self, filename1, sheetname1, filename2, sheetname2):
        if (os.path.exists(filename1) and os.path.exists(filename2)):
            wb1 = openpyxl.load_workbook(filename1)
            wb2 = openpyxl.load_workbook(filename2)
            sheets1 = wb1.sheetnames
            sheets2 = wb2.sheetnames
            if (sheetname1 in sheets1 and sheetname2 in sheets2):
                sheet1 = wb1.get_sheet_by_name(sheetname1)
                sheet2 = wb2.get_sheet_by_name(sheetname2)
                max_row = sheet1.max_row  # 最大行数
                value_dict = configKV().getKeys_values('case.ini', 'mms')
                value_list = []
                value_list.append(value_dict.get('interface_name'))
                value_list.append(value_dict.get('parameter'))
                value_list.append(value_dict.get('result'))
                value_list.append(value_dict.get('bugdescription'))
                value_list.append(value_dict.get('flag'))
                value_list.append(value_dict.get('priority'))
                # 设置报告的列标题
                sheet2.cell(1, 1, 'interface_name')
                sheet2.cell(1, 2, 'parameter')
                sheet2.cell(1, 3, 'result')
                sheet2.cell(1, 4, 'bugdescription')
                sheet2.cell(1, 5, 'flag')
                sheet2.cell(1, 6, 'time')
                sheet2.cell(1, 7, 'priority')
                sheet2.cell(1, 1).fill = styles.PatternFill(fill_type='solid', fgColor="FFFF00")
                sheet2.cell(1, 2).fill = styles.PatternFill(fill_type='solid', fgColor="FFFF00")
                sheet2.cell(1, 3).fill = styles.PatternFill(fill_type='solid', fgColor="FFFF00")
                sheet2.cell(1, 4).fill = styles.PatternFill(fill_type='solid', fgColor="FFFF00")
                sheet2.cell(1, 5).fill = styles.PatternFill(fill_type='solid', fgColor="FFFF00")
                sheet2.cell(1, 6).fill = styles.PatternFill(fill_type='solid', fgColor="FFFF00")
                sheet2.cell(1, 7).fill = styles.PatternFill(fill_type='solid', fgColor="FFFF00")
                font_cell = styles.Font(name='等线', size=14, italic=False, color=styles.colors.RED, bold=True)
                sheet2['A1'].font = font_cell
                sheet2['B1'].font = font_cell
                sheet2['C1'].font = font_cell
                sheet2['D1'].font = font_cell
                sheet2['E1'].font = font_cell
                sheet2['F1'].font = font_cell
                sheet2['G1'].font = font_cell
                # 设置行高
                sheet2.row_dimensions[1].height = 20
                # 设置列宽
                sheet2.column_dimensions['A'].width = 20
                sheet2.column_dimensions['B'].width = 50
                sheet2.column_dimensions['D'].width = 120
                sheet2.column_dimensions['E'].width = 5
                sheet2.column_dimensions['F'].width = 20

                n = 1
                for row in range(1, max_row + 1):
                    result_index = int(value_dict.get('result'))
                    result = sheet1.cell(row, result_index + 1).value
                    flag = sheet1.cell(row, int(value_dict.get('flag')) + 1).value
                    priority = sheet1.cell(row, int(value_dict.get('priority'))).value
                    # 插入数据
                    if (row > 1 and result == 0 and flag == 'Y'):
                        n = n + 1
                        interface_name = sheet1.cell(row, int(value_dict.get('interface_name')) + 1).value
                        sheet2.cell(n, 1, bytes(str(interface_name), encoding='utf-8'))
                        parameter = sheet1.cell(row, int(value_dict.get('parameter')) + 1).value
                        sheet2.cell(n, 2, bytes(str(parameter), encoding='utf-8'))
                        sheet2.cell(n, 3, bytes(str(result), encoding='utf-8'))
                        bugdescription = sheet1.cell(row, int(value_dict.get('bugdescription')) + 1).value
                        sheet2.cell(n, 4, bytes(str(bugdescription), encoding='utf-8'))
                        sheet2.cell(n, 5, bytes(str(flag), encoding='utf-8'))
                        sheet2.cell(n, 6, bytes(str(priority), encoding='utf-8'))
                wb2.save(filename2)  # 保存数据
                wb1.close()  # 关闭excel
                wb2.close()
            else:
                print(sheetname1 + ' 或 ' + sheetname2 + ' 不存在')
        else:
            print(filename1 + ' 或 ' + filename2 + ' 不存在')

    # filename:测试报告的excel的名称
    def writerReport(self, filename, datapie, datacolumn):
        now = time.strftime("%Y%m%d %H%M%S", time.localtime())
        filename2_list = filename.split('.')
        path = '../dataconfig/'
        file2 = (path + filename2_list[0] + now + '.' + filename2_list[1]).replace('\\', '/')
        workbook = xlsxwriter.Workbook(file2)
        worksheet = workbook.add_worksheet("测试总况")
        self.distribution_init(worksheet, workbook, datacolumn)
        self.overview_init(worksheet, workbook, datapie)
        workbook.close()  # workbook关闭之前不会生成excel文档


if __name__ == '__main__':
    report = reportExcel()
    # 以下两个dict（datapie，datacolumn）需要构建好数据
    datapie = {"test_name": "MMS", "test_sum": 110, "test_success": 80, "test_failed": 20, "test_not": 10,
               'defect_rate': 18, 'total_serious': 6, 'total_medium': 5, 'total_commonly': 8}
    datacolumn = {'load-exchange-rate': {'serious': 13, 'medium': 15, 'commonly': 12, 'total': 30, 'percent': 30},
                  'classify': {'serious': 13, 'medium': 5, 'commonly': 2, 'total': 30, 'percent': 20},
                  'stock-suspend': {'serious': 3, 'medium': 5, 'commonly': 12, 'total': 30, 'percent': 20},
                  'stock-brief': {'serious': 3, 'medium': 15, 'commonly': 2, 'total': 30, 'percent': 30}}
    report.writerReport('DataReport.xlsx', datapie, datacolumn)  # DataReport.xlsx为报告名称，datapie，datacolumn为报告数据